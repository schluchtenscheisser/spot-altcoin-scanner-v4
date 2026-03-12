from pathlib import Path

from openpyxl import load_workbook

from scanner.pipeline.excel_output import ExcelReportGenerator
from scanner.pipeline.output import ReportGenerator


def _row(symbol: str, **kwargs):
    row = {
        "symbol": symbol,
        "coin_name": symbol.replace("USDT", ""),
        "setup_id": "breakout_immediate_1_5d",
        "score": 80.0,
    }
    row.update(kwargs)
    return row


def _candidate(symbol: str, decision: str = "WAIT", **kwargs):
    row = {
        "rank": 1,
        "symbol": symbol,
        "coin_name": symbol.replace("USDT", ""),
        "decision": decision,
        "decision_reasons": ["entry_not_confirmed"] if decision == "WAIT" else [],
        "global_score": 80.0,
        "market_cap_usd": 1_000_000,
        "entry_ready": None,
        "risk_acceptable": None,
    }
    row.update(kwargs)
    return row


def test_json_report_contains_market_activity_fields_with_values_and_nulls() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 5}})
    rows = [
        _row("AAAUSDT", global_volume_24h_usd=100_000_000, turnover_24h=0.25, mexc_share_24h=0.05),
        _row("BBBUSDT", global_volume_24h_usd=-1, turnover_24h=float("nan"), mexc_share_24h="abc"),
    ]

    report = generator.generate_json_report(rows, [], [], [], "2026-02-28")
    first = report["setups"]["reversals"][0]
    second = report["setups"]["reversals"][1]

    assert first["global_volume_24h_usd"] == 100_000_000.0
    assert first["turnover_24h"] == 0.25
    assert first["mexc_share_24h"] == 0.05

    assert second["global_volume_24h_usd"] is None
    assert second["turnover_24h"] is None
    assert second["mexc_share_24h"] is None


def test_markdown_formats_nullable_trade_candidate_fields() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 5}})

    row = _row(
        "AAAUSDT",
        decision="WAIT",
        decision_reasons=["entry_not_confirmed"],
        global_score=80.0,
        price_usdt=100.0,
        stop_price_initial=91.9,
        analysis={"trade_levels": {"entry_trigger": 100.0}},
        slippage_bps_20k=20.5,
    )
    md = generator.generate_markdown_report([], [], [], [row], "2026-02-28")

    assert "rr_to_target_1: n/a" in md
    assert "slippage_bps_20k: 20.5000" in md


def test_excel_report_contains_trade_candidates_columns_and_null_semantics(tmp_path: Path) -> None:
    generator = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path), "top_n_per_setup": 5}})
    rows = [
        _candidate("AAAUSDT", decision="ENTER", global_score=90.0, decision_reasons=[], market_cap_usd=500_000_000),
        _candidate("BBBUSDT", decision="WAIT", rank=2, decision_reasons=["entry_not_confirmed", "btc_regime_caution"], market_cap_usd=None),
    ]

    excel_path = generator.generate_excel_report(trade_candidates=rows, run_date="2026-02-28")
    wb = load_workbook(excel_path)

    ws_trade = wb["Trade Candidates"]
    headers = [ws_trade.cell(row=1, column=i).value for i in range(1, ws_trade.max_column + 1)]
    assert headers[:5] == ["Rank", "Symbol", "Name", "Decision", "Decision Reasons"]
    market_cap_col = headers.index("Market Cap USD") + 1
    assert ws_trade.cell(row=2, column=2).value == "AAAUSDT"
    assert ws_trade.cell(row=3, column=5).value == "entry_not_confirmed | btc_regime_caution"
    assert ws_trade.cell(row=3, column=market_cap_col).value is None


def test_markdown_keeps_nullables_as_na() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 5}})

    row = _row("AAAUSDT", decision="WAIT", decision_reasons=None, global_score=80.0, rr_to_target_1=None, slippage_bps_20k=None, risk_acceptable=None)
    md = generator.generate_markdown_report([], [], [], [row], "2026-02-28")

    assert "decision_reasons: []" in md
    assert "risk_acceptable: n/a" in md
    assert "rr_to_target_1: n/a" in md
    assert "slippage_bps_20k: n/a" in md
