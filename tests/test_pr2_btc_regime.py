from pathlib import Path

from openpyxl import load_workbook

from scanner.pipeline.excel_output import ExcelReportGenerator
from scanner.pipeline.output import ReportGenerator
from scanner.pipeline.regime import compute_btc_regime_from_1d_features


def test_compute_btc_regime_risk_off_when_close_below_ema50() -> None:
    regime = compute_btc_regime_from_1d_features(
        {
            "close": 90.0,
            "ema_20": 95.0,
            "ema_50": 100.0,
        }
    )

    assert regime["state"] == "RISK_OFF"
    assert regime["checks"]["close_gt_ema50"] is False
    assert regime["checks"]["ema20_gt_ema50"] is False


def test_compute_btc_regime_risk_on_when_close_and_ema20_above_ema50() -> None:
    regime = compute_btc_regime_from_1d_features(
        {
            "close": 110.0,
            "ema_20": 105.0,
            "ema_50": 100.0,
        }
    )

    assert regime["state"] == "RISK_ON"
    assert regime["checks"]["close_gt_ema50"] is True
    assert regime["checks"]["ema20_gt_ema50"] is True


def test_markdown_summary_contains_btc_regime_state() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 1}})
    btc_regime = {
        "state": "RISK_OFF",
        "multiplier_risk_on": 1.0,
        "multiplier_risk_off": 0.85,
        "checks": {
            "close_gt_ema50": False,
            "ema20_gt_ema50": False,
        },
    }

    md = generator.generate_markdown_report([], [], [], [], "2026-02-21", btc_regime=btc_regime)

    assert "## Summary" in md
    assert "- BTC Regime: RISK_OFF" in md


def test_excel_summary_has_btc_regime_header_at_a1(tmp_path: Path) -> None:
    generator = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})
    btc_regime = {
        "state": "RISK_ON",
        "multiplier_risk_on": 1.0,
        "multiplier_risk_off": 0.85,
        "checks": {
            "close_gt_ema50": True,
            "ema20_gt_ema50": True,
        },
    }

    excel_path = generator.generate_excel_report([], [], [], [], "2026-02-21", btc_regime=btc_regime)
    wb = load_workbook(excel_path)
    ws = wb["Summary"]

    assert ws["A1"].value == "BTC Regime"
    assert ws["A2"].value == "State"
    assert ws["B2"].value == "RISK_ON"


def test_json_report_exposes_top_level_btc_regime() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 1}})
    btc_regime = {
        "state": "RISK_OFF",
        "multiplier_risk_on": 1.0,
        "multiplier_risk_off": 0.85,
        "checks": {
            "close_gt_ema50": False,
            "ema20_gt_ema50": True,
        },
    }

    report = generator.generate_json_report([], [], [], [], "2026-02-21", btc_regime=btc_regime)

    assert report["btc_regime"]["state"] == "RISK_OFF"
    assert report["btc_regime"]["checks"]["ema20_gt_ema50"] is True


def test_compute_btc_regime_neutral_when_inputs_missing() -> None:
    regime = compute_btc_regime_from_1d_features({"close": 90.0, "ema_20": 95.0})

    assert regime["state"] == "NEUTRAL"
    assert regime["checks"]["close_gt_ema50"] is None
    assert regime["checks"]["ema20_gt_ema50"] is None
