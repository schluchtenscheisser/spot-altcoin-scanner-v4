from pathlib import Path

from openpyxl import load_workbook

from scanner.pipeline.excel_output import ExcelReportGenerator
from scanner.pipeline.output import ReportGenerator


def _breakout_rows():
    return [
        {"symbol": "AAAUSDT", "setup_id": "breakout_immediate_1_5d", "score": 80.0, "final_score": 80.0, "execution_gate_pass": True, "spread_pct": 0.1, "depth_bid_0_5pct_usd": 1000, "depth_ask_0_5pct_usd": 1000, "depth_bid_1pct_usd": 2000, "depth_ask_1pct_usd": 2000},
        {"symbol": "BBBUSDT", "setup_id": "breakout_retest_1_5d", "score": 82.0, "final_score": 82.0, "execution_gate_pass": False, "spread_pct": 0.2, "depth_bid_0_5pct_usd": 900, "depth_ask_0_5pct_usd": 900, "depth_bid_1pct_usd": 1800, "depth_ask_1pct_usd": 1800},
    ]


def test_pr5_markdown_contains_sot_sections() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 10}})

    md = generator.generate_markdown_report([], _breakout_rows(), [], _breakout_rows(), "2026-02-22")

    assert "## ENTER Candidates" in md
    assert "## WAIT Candidates" in md


def test_pr5_json_contains_breakout_setup_lists() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 10}})

    report = generator.generate_json_report([], _breakout_rows(), [], [], "2026-02-22")

    assert report["setups"]["breakout_immediate_1_5d"][0]["setup_id"] == "breakout_immediate_1_5d"
    assert report["setups"]["breakout_retest_1_5d"][0]["setup_id"] == "breakout_retest_1_5d"


def test_pr5_excel_has_legacy_and_new_breakout_sheets(tmp_path: Path) -> None:
    generator = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    excel_path = generator.generate_excel_report([], _breakout_rows(), [], [], "2026-02-22")
    wb = load_workbook(excel_path)

    assert "Breakout Setups" in wb.sheetnames
    assert "Breakout Immediate 1-5D" in wb.sheetnames
    assert "Breakout Retest 1-5D" in wb.sheetnames


def test_pr5_markdown_renders_wait_reasons_from_trade_candidates() -> None:
    generator = ReportGenerator({"output": {"top_n_per_setup": 10}})

    rows = _breakout_rows()
    rows[0]["decision"] = "WAIT"
    rows[0]["decision_reasons"] = ["entry_not_confirmed", "btc_regime_caution"]
    rows[0]["global_score"] = 80.0

    md = generator.generate_markdown_report([], rows, [], rows, "2026-02-22")

    assert "decision_reasons: entry_not_confirmed, btc_regime_caution" in md


def test_pr5_excel_setup_sheet_contains_execution_columns(tmp_path: Path) -> None:
    generator = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    excel_path = generator.generate_excel_report([], _breakout_rows(), [], [], "2026-02-22")
    wb = load_workbook(excel_path)
    ws = wb["Breakout Immediate 1-5D"]

    headers = [ws.cell(row=1, column=idx).value for idx in range(1, 12)]
    assert headers[4] == "Execution Gate Pass"
    assert headers[5] == "Spread %"
    assert headers[8] == "Depth Bid 1.0% USD"
