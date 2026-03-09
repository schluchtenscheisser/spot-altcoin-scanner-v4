from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from scanner.pipeline.excel_output import ExcelReportGenerator
from scanner.pipeline.output import ReportGenerator


RUN_DATE = "2026-03-08"


def _global_top20_fixture() -> list[dict]:
    return [
        {
            "symbol": "BBBUSDT",
            "coin_name": "Beta",
            "decision": "WAIT",
            "decision_reasons": ["entry_not_confirmed", "überhitzung"],
            "global_score": 88.0,
            "risk_acceptable": None,
            "rr_to_tp10": None,
            "tradeability_class": "UNKNOWN",
            "spread_pct": 0.0012,
            "depth_bid_1pct_usd": 12000.0,
            "depth_ask_1pct_usd": 13000.0,
            "slippage_bps_20k": 18.5,
            "best_setup_type": "breakout",
        },
        {
            "symbol": "AAAUSDT",
            "coin_name": "Alpha",
            "decision": "ENTER",
            "decision_reasons": [],
            "global_score": 88.0,
            "risk_acceptable": True,
            "rr_to_tp10": 1.8,
            "tradeability_class": "A",
            "spread_pct": 0.0009,
            "depth_bid_1pct_usd": 25000.0,
            "depth_ask_1pct_usd": 26000.0,
            "slippage_bps_20k": 9.2,
            "best_setup_type": "pullback",
        },
        {
            "symbol": "CCCUSDT",
            "coin_name": "Gamma",
            "decision": "NO_TRADE",
            "decision_reasons": ["risk_budget_exceeded"],
            "global_score": 91.0,
            "risk_acceptable": False,
            "rr_to_tp10": 0.5,
            "tradeability_class": "B",
        },
    ]


def _sheet_rows_by_symbol(excel_path: Path, sheet_name: str = "Trade Candidates") -> dict[str, dict[str, object]]:
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    by_symbol: dict[str, dict[str, object]] = {}
    for row_idx in range(2, ws.max_row + 1):
        row = {headers[col_idx - 1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)}
        by_symbol[str(row["Symbol"])] = row
    return by_symbol


def _sheet_symbols_in_order(excel_path: Path, sheet_name: str = "Trade Candidates") -> list[str]:
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    return [str(ws.cell(row=row_idx, column=2).value) for row_idx in range(2, ws.max_row + 1)]


def test_pr22_json_markdown_excel_semantics_are_consistent(tmp_path: Path) -> None:
    report_gen = ReportGenerator({"output": {"reports_dir": str(tmp_path)}})
    excel_gen = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    global_top20 = _global_top20_fixture()
    json_report = report_gen.generate_json_report([], [], [], global_top20, RUN_DATE)
    markdown = report_gen.generate_markdown_report([], [], [], global_top20, RUN_DATE)
    excel_path = excel_gen.generate_excel_report(json_report["trade_candidates"], RUN_DATE)

    json_symbols = [row["symbol"] for row in json_report["trade_candidates"]]
    assert json_symbols == ["AAAUSDT", "BBBUSDT", "CCCUSDT"]
    assert _sheet_symbols_in_order(excel_path) == json_symbols

    assert markdown.index("### 1. AAAUSDT (Alpha)") < markdown.index("### 2. BBBUSDT (Beta)")
    assert "decision_reasons: entry_not_confirmed, überhitzung" in markdown
    assert "- risk_acceptable: n/a" in markdown
    assert "- rr_to_tp10: n/a" in markdown

    excel_rows = _sheet_rows_by_symbol(excel_path)
    bbb = excel_rows["BBBUSDT"]
    assert bbb["Decision"] == "WAIT"
    assert bbb["Decision Reasons"] == "entry_not_confirmed | überhitzung"
    assert bbb["Risk Acceptable"] is None
    assert bbb["RR to TP10"] is None
    assert bbb["Tradeability Class"] == "UNKNOWN"


def test_pr22_markdown_and_excel_fail_on_missing_required_trade_candidate_fields(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    report_gen = ReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    def _invalid_json(*args, **kwargs):
        return {
            "trade_candidates": [{"rank": 1, "symbol": "AAAUSDT", "decision": "ENTER"}],
            "run_manifest": {},
            "btc_regime": {},
        }

    monkeypatch.setattr(report_gen, "generate_json_report", _invalid_json)
    with pytest.raises(ValueError, match="missing required fields"):
        report_gen.generate_markdown_report([], [], [], _global_top20_fixture(), RUN_DATE)

    excel_gen = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})
    with pytest.raises(ValueError, match="missing required fields"):
        excel_gen.generate_excel_report([{"rank": 1, "symbol": "AAAUSDT", "decision": "ENTER"}], RUN_DATE)


def test_pr22_excel_preflight_is_atomic_no_file_on_invalid_input(tmp_path: Path) -> None:
    excel_gen = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})
    target = tmp_path / f"{RUN_DATE}.xlsx"

    with pytest.raises(ValueError, match="decision_reasons"):
        excel_gen.generate_excel_report(
            [{"rank": 1, "symbol": "AAAUSDT", "decision": "ENTER", "decision_reasons": "bad"}],
            RUN_DATE,
        )

    assert not target.exists()


def test_pr22_empty_trade_candidates_render_deterministically(tmp_path: Path) -> None:
    report_gen = ReportGenerator({"output": {"reports_dir": str(tmp_path)}})
    excel_gen = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    md_first = report_gen.generate_markdown_report([], [], [], [], RUN_DATE)
    md_second = report_gen.generate_markdown_report([], [], [], [], RUN_DATE)
    assert md_first == md_second
    assert "*No ENTER candidates.*" in md_first
    assert "*No WAIT candidates.*" in md_first

    excel_path = excel_gen.generate_excel_report([], RUN_DATE)
    wb = load_workbook(excel_path)
    assert wb["Trade Candidates"].max_row == 1
    assert wb["Enter Candidates"].max_row == 1
    assert wb["Wait Candidates"].max_row == 1


def test_pr22_deterministic_ordering_with_tied_scores_across_formats(tmp_path: Path) -> None:
    report_gen = ReportGenerator({"output": {"reports_dir": str(tmp_path)}})
    excel_gen = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    global_top20 = [
        {
            "symbol": "ZZZUSDT",
            "coin_name": "ZZZ",
            "decision": "ENTER",
            "decision_reasons": [],
            "global_score": 90.0,
            "best_setup_type": "b",
        },
        {
            "symbol": "AAAUSDT",
            "coin_name": "AAA",
            "decision": "ENTER",
            "decision_reasons": [],
            "global_score": 90.0,
            "best_setup_type": "a",
        },
    ]

    json_first = report_gen.generate_json_report([], [], [], global_top20, RUN_DATE)
    json_second = report_gen.generate_json_report([], [], [], global_top20, RUN_DATE)
    assert [row["symbol"] for row in json_first["trade_candidates"]] == ["AAAUSDT", "ZZZUSDT"]
    assert [row["symbol"] for row in json_first["trade_candidates"]] == [row["symbol"] for row in json_second["trade_candidates"]]

    md_first = report_gen.generate_markdown_report([], [], [], global_top20, RUN_DATE)
    md_second = report_gen.generate_markdown_report([], [], [], global_top20, RUN_DATE)
    assert md_first == md_second
    assert md_first.index("### 1. AAAUSDT (AAA)") < md_first.index("### 2. ZZZUSDT (ZZZ)")

    excel_first = excel_gen.generate_excel_report(json_first["trade_candidates"], RUN_DATE)
    excel_second = excel_gen.generate_excel_report(json_second["trade_candidates"], RUN_DATE)
    assert _sheet_symbols_in_order(excel_first) == ["AAAUSDT", "ZZZUSDT"]
    assert _sheet_symbols_in_order(excel_second) == ["AAAUSDT", "ZZZUSDT"]


def test_pr22_excel_summary_contains_entry_state_counts(tmp_path: Path) -> None:
    report_gen = ReportGenerator({"output": {"reports_dir": str(tmp_path)}})
    excel_gen = ExcelReportGenerator({"output": {"reports_dir": str(tmp_path)}})

    json_report = report_gen.generate_json_report([], [], [], _global_top20_fixture(), RUN_DATE)
    excel_path = excel_gen.generate_excel_report(json_report["trade_candidates"], RUN_DATE)

    wb = load_workbook(excel_path)
    ws = wb["Summary"]
    values = {
        ws.cell(row=row_idx, column=1).value: ws.cell(row=row_idx, column=2).value
        for row_idx in range(1, ws.max_row + 1)
        if ws.cell(row=row_idx, column=1).value is not None
    }

    assert values["entry_state_counts_all.early"] == 0
    assert values["entry_state_counts_all.at_trigger"] == 0
    assert values["entry_state_counts_all.late"] == 0
    assert values["entry_state_counts_all.chased"] == 0
    assert values["entry_state_counts_all.null"] == 3
