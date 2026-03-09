"""
Excel Output Generation
=======================

Generates Excel workbooks from canonical trade_candidates SoT.
"""

import logging
import math
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generates Excel reports as pure SoT views."""

    REQUIRED_FIELDS = {"rank", "symbol", "decision", "decision_reasons"}
    ALLOWED_DECISIONS = {"ENTER", "WAIT", "NO_TRADE"}
    CANDIDATE_COLUMNS = [
        ("rank", "Rank"),
        ("symbol", "Symbol"),
        ("coin_name", "Name"),
        ("decision", "Decision"),
        ("decision_reasons", "Decision Reasons"),
        ("best_setup_type", "Best Setup"),
        ("setup_subtype", "Setup Subtype"),
        ("global_score", "Global Score"),
        ("setup_score", "Setup Score"),
        ("entry_ready", "Entry Ready"),
        ("entry_readiness_reasons", "Entry Readiness Reasons"),
        ("execution_mode", "Execution Mode"),
        ("tradeability_class", "Tradeability Class"),
        ("risk_acceptable", "Risk Acceptable"),
        ("entry_price_usdt", "Entry Price (USDT)"),
        ("current_price_usdt", "Current Price (USDT)"),
        ("distance_to_entry_pct", "Distance to Entry (%)"),
        ("entry_state", "Entry State"),
        ("stop_price_initial", "Stop Price Initial"),
        ("risk_pct_to_stop", "Risk % to Stop"),
        ("tp10_price", "TP10 Price"),
        ("tp20_price", "TP20 Price"),
        ("rr_to_tp10", "RR to TP10"),
        ("rr_to_tp20", "RR to TP20"),
        ("spread_pct", "Spread %"),
        ("depth_bid_1pct_usd", "Depth Bid 1.0% USD"),
        ("depth_ask_1pct_usd", "Depth Ask 1.0% USD"),
        ("slippage_bps_5k", "Slippage BPS 5k"),
        ("slippage_bps_20k", "Slippage BPS 20k"),
        ("market_cap_usd", "Market Cap USD"),
        ("btc_regime", "BTC Regime"),
        ("flags", "Flags"),
    ]

    def __init__(self, config: Dict[str, Any]):
        if hasattr(config, "raw"):
            output_config = config.raw.get("output", {})
        else:
            output_config = config.get("output", {})

        self.reports_dir = Path(output_config.get("reports_dir", "reports"))
        self.reports_dir.mkdir(parents=True, exist_ok=True)

        logger.info("Excel Report Generator initialized: reports_dir=%s", self.reports_dir)

    def generate_excel_report(
        self,
        trade_candidates: List[Dict[str, Any]],
        run_date: str,
        run_manifest: Dict[str, Any] | None = None,
        metadata: Dict[str, Any] | None = None,
        btc_regime: Dict[str, Any] | None = None,
    ) -> Path:
        """Generate workbook from canonical trade_candidates SoT."""
        logger.info("Generating Excel report for %s", run_date)

        validated = self._validate_trade_candidates(trade_candidates)
        enter_candidates = [row for row in validated if row.get("decision") == "ENTER"]
        wait_candidates = [row for row in validated if row.get("decision") == "WAIT"]

        wb = Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(
            wb,
            run_date=run_date,
            trade_candidates=validated,
            run_manifest=run_manifest or {},
            metadata=metadata,
            btc_regime=btc_regime,
        )
        self._create_trade_candidates_sheet(wb, "Trade Candidates", validated)
        self._create_trade_candidates_sheet(wb, "Enter Candidates", enter_candidates)
        self._create_trade_candidates_sheet(wb, "Wait Candidates", wait_candidates)

        excel_path = self.reports_dir / f"{run_date}.xlsx"
        wb.save(excel_path)
        logger.info("Excel report saved: %s", excel_path)
        return excel_path

    def _validate_trade_candidates(self, trade_candidates: Any) -> List[Dict[str, Any]]:
        if not isinstance(trade_candidates, list):
            raise ValueError("trade_candidates must be a list")

        validated: List[Dict[str, Any]] = []
        for idx, row in enumerate(trade_candidates):
            if not isinstance(row, dict):
                raise ValueError(f"trade_candidates[{idx}] must be an object")

            missing = [field for field in self.REQUIRED_FIELDS if field not in row]
            if missing:
                raise ValueError(f"trade_candidates[{idx}] missing required fields: {', '.join(missing)}")

            decision = row.get("decision")
            if decision not in self.ALLOWED_DECISIONS:
                raise ValueError(f"trade_candidates[{idx}].decision invalid: {decision}")

            reasons = row.get("decision_reasons")
            if reasons is not None and not isinstance(reasons, list):
                raise ValueError(f"trade_candidates[{idx}].decision_reasons must be list or null")

            validated.append(row)
        return validated

    def _create_summary_sheet(
        self,
        wb: Workbook,
        run_date: str,
        trade_candidates: List[Dict[str, Any]],
        run_manifest: Dict[str, Any],
        metadata: Dict[str, Any] | None,
        btc_regime: Dict[str, Any] | None,
    ) -> None:
        ws = wb.create_sheet("Summary", 0)

        btc_regime = btc_regime or {}
        checks = btc_regime.get("checks") or {}
        counts = {"ENTER": 0, "WAIT": 0, "NO_TRADE": 0}
        for row in trade_candidates:
            decision = row.get("decision")
            if decision in counts:
                counts[decision] += 1

        ws["A1"] = "BTC Regime"
        ws["A2"] = "State"
        ws["B2"] = btc_regime.get("state", "RISK_OFF")
        ws["A3"] = "Multiplier (Risk-On)"
        ws["B3"] = self._sanitize_float_or_none(btc_regime.get("multiplier_risk_on"), default=1.0)
        ws["A4"] = "Multiplier (Risk-Off)"
        ws["B4"] = self._sanitize_float_or_none(btc_regime.get("multiplier_risk_off"), default=0.85)
        ws["A5"] = "close>ema50"
        ws["B5"] = checks.get("close_gt_ema50")
        ws["A6"] = "ema20>ema50"
        ws["B6"] = checks.get("ema20_gt_ema50")

        ws["A8"] = "Metric"
        ws["B8"] = "Value"
        for cell in ["A8", "B8"]:
            ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal="center")

        row_idx = 9
        entry_state_order = ("early", "at_trigger", "late", "chased", "null")
        entry_state_counts = {state: 0 for state in entry_state_order}
        for row in trade_candidates:
            entry_state = row.get("entry_state")
            key = entry_state if entry_state in {"early", "at_trigger", "late", "chased"} else "null"
            entry_state_counts[key] += 1

        summary_rows = [
            ("Run Date", run_date),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Trade Candidates", len(trade_candidates)),
            ("ENTER Candidates", counts["ENTER"]),
            ("WAIT Candidates", counts["WAIT"]),
            ("NO_TRADE Candidates", counts["NO_TRADE"]),
            ("entry_state_counts_all.early", entry_state_counts["early"]),
            ("entry_state_counts_all.at_trigger", entry_state_counts["at_trigger"]),
            ("entry_state_counts_all.late", entry_state_counts["late"]),
            ("entry_state_counts_all.chased", entry_state_counts["chased"]),
            ("entry_state_counts_all.null", entry_state_counts["null"]),
            ("run_id", run_manifest.get("run_id")),
            ("canonical_schema_version", run_manifest.get("canonical_schema_version")),
        ]

        if metadata:
            summary_rows.extend(
                [
                    ("Total Symbols Scanned", metadata.get("universe_size")),
                    ("Symbols Filtered (MidCaps)", metadata.get("filtered_size")),
                    ("Symbols in Shortlist", metadata.get("shortlist_size")),
                ]
            )

        for key, value in summary_rows:
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=value)
            row_idx += 1

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 28

    def _create_trade_candidates_sheet(self, wb: Workbook, sheet_name: str, rows: List[Dict[str, Any]]) -> None:
        ws = wb.create_sheet(sheet_name)

        for col_idx, (_, header) in enumerate(self.CANDIDATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, (key, _) in enumerate(self.CANDIDATE_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=self._to_excel_cell_value(row.get(key)))

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col_idx, (_, header) in enumerate(self.CANDIDATE_COLUMNS, start=1):
            col = get_column_letter(col_idx)
            ws.column_dimensions[col].width = max(12, min(len(header) + 4, 36))

    @classmethod
    def _to_excel_cell_value(cls, value: Any) -> Any:
        if isinstance(value, list):
            return " | ".join(str(item) for item in value)
        if isinstance(value, dict):
            return str(value)
        return cls._sanitize_float_if_needed(value)

    @staticmethod
    def _sanitize_float_if_needed(value: Any) -> Any:
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return value

    @staticmethod
    def _sanitize_float_or_none(value: Any, default: float | None = None) -> float | None:
        if value is None:
            return default
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return default
        if math.isnan(numeric) or math.isinf(numeric):
            return default
        return numeric
